from django.shortcuts import render
from django.views import generic
from django.urls import reverse_lazy

from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.decorators import login_required, permission_required

from .models import Cliente, TipoPapel, Gramaje, Papel, Formato_Libro, Modo, Contacto, Libro
from .forms import ClienteForm, TipoPapelForm, GramajeForm, PapelForm, Formato_LibroForm, ModoForm, ContactoForm, LibroForm

from bases.views import SinPrivilegios
from django.http import HttpResponse

#CLIENTES
class ClienteView(SinPrivilegios, generic.ListView):
    permission_required = "bd.view_cliente"
    model = Cliente
    template_name = "bd/cliente_list.html"
    context_object_name = "obj"

class ClienteNew(SuccessMessageMixin, SinPrivilegios, generic.CreateView):
    model = Cliente
    template_name = "bd/cliente_form.html"
    context_object_name = "obj"
    form_class = ClienteForm
    success_url = reverse_lazy("bd:cliente_list")
    success_message = "Cliente Creado"
    permission_required ="bd.add_cliente"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class ClienteEdit(SuccessMessageMixin, SinPrivilegios, generic.UpdateView):
    model = Cliente
    template_name = "bd/cliente_form.html"
    context_object_name = "obj"
    form_class = ClienteForm
    success_url = reverse_lazy("bd:cliente_list")
    success_message = "Cliente Actualizado Correctamente"
    permission_required = "bd.change_cliente"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

#TIPOS DE PAPEL
class TipoPapelView(SinPrivilegios, generic.ListView):
    permission_required = "bd.view_tipopapel"
    model = TipoPapel
    template_name = "bd/tipopapel_list.html"
    context_object_name = "obj"

class TipoPapelNew(SuccessMessageMixin, SinPrivilegios, generic.CreateView):
    model = TipoPapel
    template_name = "bd/tipopapel_form.html"
    context_object_name = "obj"
    form_class = TipoPapelForm
    success_url = reverse_lazy("bd:tipopapel_list")
    success_message = "Tipo de Papel Creado"
    permission_required ="bd.add_tipopapel"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class TipoPapelEdit(SuccessMessageMixin, SinPrivilegios, generic.UpdateView):
    model = TipoPapel
    template_name = "bd/tipopapel_form.html"
    context_object_name = "obj"
    form_class = TipoPapelForm
    success_url = reverse_lazy("bd:tipopapel_list")
    success_message = "Tipo de Papel Actualizado Correctamente"
    permission_required = "bd.change_tipopapel"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

#GRAMAJES
class GramajeView(SinPrivilegios, generic.ListView):
    permission_required = "bd.view_gramaje"
    model = Gramaje
    template_name = "bd/gramaje_list.html"
    context_object_name = "obj"

class GramajeNew(SuccessMessageMixin, SinPrivilegios, generic.CreateView):
    model = Gramaje
    template_name = "bd/gramaje_form.html"
    context_object_name = "obj"
    form_class = GramajeForm
    success_url = reverse_lazy("bd:gramaje_list")
    success_message = "Tipo de Papel Creado"
    permission_required ="bd.add_gramaje"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class GramajeEdit(SuccessMessageMixin, SinPrivilegios, generic.UpdateView):
    model = Gramaje
    template_name = "bd/gramaje_form.html"
    context_object_name = "obj"
    form_class = GramajeForm
    success_url = reverse_lazy("bd:gramaje_list")
    success_message = "Tipo de Papel Actualizado Correctamente"
    permission_required = "bd.change_gramaje"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

#PAPELES
class PapelView(SinPrivilegios, generic.ListView):
    permission_required = "bd.view_papel"
    model = Papel
    template_name = "bd/papel_list.html"
    context_object_name = "obj"

class PapelNew(SuccessMessageMixin, SinPrivilegios, generic.CreateView):
    model = Papel
    template_name = "bd/papel_form.html"
    context_object_name = "obj"
    form_class = PapelForm
    success_url = reverse_lazy("bd:papel_list")
    success_message = "Tipo de Papel Creado"
    permission_required ="bd.add_papel"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class PapelEdit(SuccessMessageMixin, SinPrivilegios, generic.UpdateView):
    model = Papel
    template_name = "bd/papel_form.html"
    context_object_name = "obj"
    form_class = PapelForm
    success_url = reverse_lazy("bd:papel_list")
    success_message = "Tipo de Papel Actualizado Correctamente"
    permission_required = "bd.change_papel"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

#FORMATO
class Formato_LibroView(SinPrivilegios, generic.ListView):
    permission_required = "bd.view_formato_libro"
    model = Formato_Libro
    template_name = "bd/formato_libro_list.html"
    context_object_name = "obj"

class Formato_LibroNew(SuccessMessageMixin, SinPrivilegios, generic.CreateView):
    model = Formato_Libro
    template_name = "bd/formato_libro_form.html"
    context_object_name = "obj"
    form_class = Formato_LibroForm
    success_url = reverse_lazy("bd:formato_libro_list")
    success_message = "Formato Creado"
    permission_required ="bd.add_formato_libro"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class Formato_LibroEdit(SuccessMessageMixin, SinPrivilegios, generic.UpdateView):
    model = Formato_Libro
    template_name = "bd/formato_libro_form.html"
    context_object_name = "obj"
    form_class = Formato_LibroForm
    success_url = reverse_lazy("bd:formato_libro_list")
    success_message = "Formato Actualizado Correctamente"
    permission_required = "bd.change_cliente"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)


#MODO
class ModoView(SinPrivilegios, generic.ListView):
    permission_required = "bd.view_modo"
    model = Modo
    template_name = "bd/modo_list.html"
    context_object_name = "obj"

class ModoNew(SuccessMessageMixin, SinPrivilegios, generic.CreateView):
    model = Modo
    template_name = "bd/modo_form.html"
    context_object_name = "obj"
    form_class = ModoForm
    success_url = reverse_lazy("bd:modo_list")
    success_message = "Formato Creado"
    permission_required ="bd.add_modo"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class ModoEdit(SuccessMessageMixin, SinPrivilegios, generic.UpdateView):
    model = Modo
    template_name = "bd/modo_form.html"
    context_object_name = "obj"
    form_class = ModoForm
    success_url = reverse_lazy("bd:modo_list")
    success_message = "Formato Actualizado Correctamente"
    permission_required = "bd.change_cliente"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

#CLIENTES
class ContactoView(SinPrivilegios, generic.ListView):
    permission_required = "bd.view_contacto"
    model = Contacto
    template_name = "bd/contacto_list.html"
    context_object_name = "obj"

class ContactoNew(SuccessMessageMixin, SinPrivilegios, generic.CreateView):
    model = Contacto
    template_name = "bd/contacto_form.html"
    context_object_name = "obj"
    form_class = ContactoForm
    success_url = reverse_lazy("bd:contacto_list")
    success_message = "Contacto Creado"
    permission_required ="bd.add_contacto"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class ContactoEdit(SuccessMessageMixin, SinPrivilegios, generic.UpdateView):
    model = Contacto
    template_name = "bd/contacto_form.html"
    context_object_name = "obj"
    form_class = ContactoForm
    success_url = reverse_lazy("bd:contacto_list")
    success_message = "Contacto Actualizado Correctamente"
    permission_required = "bd.change_contacto"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

#LIBROS
class LibroView(SinPrivilegios, generic.ListView):
    permission_required = "bd.view_libro"
    model = Libro
    template_name = "bd/libro_list.html"
    context_object_name = "obj"

class LibroNew(SuccessMessageMixin, SinPrivilegios, generic.CreateView):
    model = Libro
    template_name = "bd/libro_form.html"
    context_object_name = "obj"
    form_class = LibroForm
    success_url = reverse_lazy("bd:libro_list")
    success_message = "Libro Creado"
    permission_required ="bd.create_libro"

    def form_valid(self, form):
        
        form.instance.uc = self.request.user
        return super().form_valid(form)

class LibroEdit(SuccessMessageMixin, SinPrivilegios, generic.UpdateView):
    model = Libro
    template_name = "bd/libro_form.html"
    context_object_name = "obj"
    form_class = LibroForm
    success_url = reverse_lazy("bd:libro_list")
    success_message = "Libro Actualizado Correctamente"
    permission_required = "bd.change_libro"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

#INACTIVACIONES
@login_required(login_url = '/login/')
@permission_required('bd.change_cliente', login_url= 'bases:sin_privilegios')
def cliente_inactivar(request, id):
    template_name = "bd/cliente_inactivar.html"
    contexto={}
    cliente = Cliente.objects.filter(pk=id).first()
    
    if not cliente:
        return HttpResponse('Cliente no existe ' + str(id))

    if request.method == 'GET':
        contexto = {'obj':cliente}
    
    if request.method == 'POST':
        if cliente.estado:
            cliente.estado = False
            cliente.save()
            contexto = {'obj':'OK'}
            
            return HttpResponse('Cliente Inactivado')
            
        else:
            cliente.estado = True
            cliente.save()
            contexto = {'obj':'OK'}
            return HttpResponse('Cliente Activado')
    return render(request, template_name,contexto)

@login_required(login_url = '/login/')
@permission_required('bd.change_tipopapel', login_url= 'bases:sin_privilegios')
def tipopapel_inactivar(request, id):
    template_name = "bd/tipopapel_inactivar.html"
    contexto={}
    tipopapel = TipoPapel.objects.filter(pk=id).first()
    
    if not tipopapel:
        return HttpResponse('Tipo de Papel no existe ' + str(id))

    if request.method == 'GET':
        contexto = {'obj':tipopapel}
    
    if request.method == 'POST':
        if tipopapel.estado:
            tipopapel.estado = False
            tipopapel.save()
            contexto = {'obj':'OK'}
            
            return HttpResponse('Tipo de Papel Inactivado')
            
        else:
            tipopapel.estado = True
            tipopapel.save()
            contexto = {'obj':'OK'}
            return HttpResponse('Tipo de Papel Activado')
    return render(request, template_name,contexto)

@login_required(login_url = '/login/')
@permission_required('bd.change_gramaje', login_url= 'bases:sin_privilegios')
def gramaje_inactivar(request, id):
    template_name = "bd/gramaje_inactivar.html"
    contexto={}
    gramaje = Gramaje.objects.filter(pk=id).first()
    
    if not gramaje:
        return HttpResponse('Gramaje no existe ' + str(id))

    if request.method == 'GET':
        contexto = {'obj':gramaje}
    
    if request.method == 'POST':
        if gramaje.estado:
            gramaje.estado = False
            gramaje.save()
            contexto = {'obj':'OK'}
            
            return HttpResponse('Gramaje Inactivado')
            
        else:
            gramaje.estado = True
            gramaje.save()
            contexto = {'obj':'OK'}
            return HttpResponse('Gramaje Activado')
    return render(request, template_name,contexto)

@login_required(login_url = '/login/')
@permission_required('bd.change_gramaje', login_url= 'bases:sin_privilegios')
def papel_inactivar(request, id):
    template_name = "bd/papel_inactivar.html"
    contexto={}
    papel = Papel.objects.filter(pk=id).first()
    
    if not papel:
        return HttpResponse('Papel no existe ' + str(id))

    if request.method == 'GET':
        contexto = {'obj':papel}
    
    if request.method == 'POST':
        if papel.estado:
            papel.estado = False
            papel.save()
            contexto = {'obj':'OK'}
            
            return HttpResponse('Papel Inactivado')
            
        else:
            papel.estado = True
            papel.save()
            contexto = {'obj':'OK'}
            return HttpResponse('Papel Activado')
    return render(request, template_name,contexto)

@login_required(login_url = '/login/')
@permission_required('bd.change_formato_libro', login_url= 'bases:sin_privilegios')
def formato_libro_inactivar(request, id):
    template_name = "bd/formato_libro_inactivar.html"
    contexto={}
    formato_libro = Formato_Libro.objects.filter(pk=id).first()
    
    if not formato_libro:
        return HttpResponse('Formato no existe ' + str(id))

    if request.method == 'GET':
        contexto = {'obj':formato_libro}
    
    if request.method == 'POST':
        if formato_libro.estado:
            formato_libro.estado = False
            formato_libro.save()
            contexto = {'obj':'OK'}
            
            return HttpResponse('Formato Inactivado')
            
        else:
            formato_libro.estado = True
            formato_libro.save()
            contexto = {'obj':'OK'}
            return HttpResponse('Formato Activado')
    return render(request, template_name,contexto)


@login_required(login_url = '/login/')
@permission_required('bd.change_modo', login_url= 'bases:sin_privilegios')
def modo_inactivar(request, id):
    template_name = "bd/modo_inactivar.html"
    contexto={}
    modo = Modo.objects.filter(pk=id).first()
    
    if not modo:
        return HttpResponse('Modo no existe ' + str(id))

    if request.method == 'GET':
        contexto = {'obj':modo}
    
    if request.method == 'POST':
        if modo.estado:
            modo.estado = False
            modo.save()
            contexto = {'obj':'OK'}
            
            return HttpResponse('Modo Inactivado')
            
        else:
            modo.estado = True
            modo.save()
            contexto = {'obj':'OK'}
            return HttpResponse('Modo Activado')
    return render(request, template_name,contexto)

#INACTIVACIONES
@login_required(login_url = '/login/')
@permission_required('bd.change_contacto', login_url= 'bases:sin_privilegios')
def contacto_inactivar(request, id):
    template_name = "bd/contacto_inactivar.html"
    contexto={}
    contacto = Contacto.objects.filter(pk=id).first()
    
    if not contacto:
        return HttpResponse('Contacto no existe ' + str(id))

    if request.method == 'GET':
        contexto = {'obj':contacto}
    
    if request.method == 'POST':
        if contacto.estado:
            contacto.estado = False
            contacto.save()
            contexto = {'obj':'OK'}
            
            return HttpResponse('Contacto Inactivado')
            
        else:
            contacto.estado = True
            contacto.save()
            contexto = {'obj':'OK'}
            return HttpResponse('Contacto Activado')
    return render(request, template_name,contexto)

@login_required(login_url = '/login/')
@permission_required('bd.change_libro', login_url= 'bases:sin_privilegios')
def libro_inactivar(request, id):
    template_name = "bd/libro_inactivar.html"
    contexto={}
    libro = Libro.objects.filter(pk=id).first()
    
    if not libro:
        return HttpResponse('Libro no existe ' + str(id))

    if request.method == 'GET':
        contexto = {'obj':libro}
    
    if request.method == 'POST':
        if libro.estado:
            libro.estado = False
            libro.save()
            contexto = {'obj':'OK'}
            
            return HttpResponse('Libro Inactivado')
            
        else:
            libro.estado = True
            libro.save()
            contexto = {'obj':'OK'}
            return HttpResponse('Libro Activado')
    return render(request, template_name,contexto)

def libros_importacion(request):
    import openpyxl
    import psycopg2
    import datetime

    conn = psycopg2.connect("host='localhost' port='5432' dbname='db_servinform' user='carlos' password='1234'")
    cur = conn.cursor()

    doc = openpyxl.load_workbook('muestra2.xlsm')

    hojas = doc.sheetnames

    hoja = doc['Sheet1']
    contador = 1
    contador2 = 1

    col_isbn = ''
    col_paginas = ''
    col_editorial = ''
    col_datovariable = ''
    col_titulo = ''
    col_solapas = ''
    col_codigo = ''
    col_bitono = ''
    col_color = ''
    col_modo = ''
    col_papel = ''
    col_gramaje = ''
    col_tamaño = ''
    col_estado = ''

    for fila in hoja.rows:
    
        if contador == 1:
            if hoja.cell(row=contador, column = contador2).value == 'ISBN':
                col_isbn = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Páginas':
                col_paginas = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Editorial':
                col_editorial = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Dato variable':
                col_datovariable = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Título':
                col_titulo = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Solapas':
                col_solapas = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Código':
                col_codigo = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Páginas bitono':
                col_bitono = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Páginas color':
                col_color = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Tipo de papel':
                col_papel = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Gramaje':
                col_gramaje = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Formato del libro':
                col_tamaño = contador2
            elif hoja.cell(row=contador, column = contador2).value == 'Estado':
                col_estado = contador2
            
            contador2 = contador2 + 1
        


    for fila in hoja.rows:
        
        if contador == 1:
            pass
        # elif contador >52:
        #     break
        
        else:
            try:
                isbn = hoja.cell(row=contador, column = col_isbn).value
                paginas = hoja.cell(row=contador, column = col_paginas).value
                if paginas == '#N/A':
                    paginas = 0
                try:
                    editorial = hoja.cell(row=contador, column = col_editorial).value 
                except:
                    editorial = 'EDITORIAL SINTESIS, S.A.'
                datovariable = hoja.cell(row=contador, column = col_datovariable).value
                if datovariable.upper() == "SI":
                    datovariable = True
                else:
                    datovariable = False    

                estado = hoja.cell(row=contador, column = col_estado).value
                if (estado.upper() == 'OK' or estado == ''):
                    estado = True
                else:
                    estado = False
                cod_interno = hoja.cell(row=contador, column = col_codigo).value
                titulo = hoja.cell(row=contador, column = col_titulo).value
                bitono = hoja.cell(row=contador, column = col_bitono).value
                if (bitono == '' or bitono == None):
                    bitono = 0
                
                color = hoja.cell(row=contador, column = col_color).value
                if (color == '' or color == None):
                    color = 0

                if (bitono == 0 and color == 0): 
                    modo = "B/N"
                elif (bitono != 0 and color != 0):
                    insercion = int(bitono) + int(color)
                    if insercion == paginas:
                        modo = 'Color'
                    else:
                        modo = 'Hibrido'
                    

                papel = hoja.cell(row=contador, column = col_papel).value
            
                gramaje = hoja.cell(row=contador, column = col_gramaje).value
                tamaño = hoja.cell(row=contador, column = col_tamaño).value
                solapas = hoja.cell(row=contador, column = col_solapas).value
                if solapas.upper() == 'NO':
                    solapas = False
                else:
                    solapas = True
                contador_impresion = 0
                fc = datetime.datetime.now()
                fm = fc
                uc_id = 1
                print(cod_interno)
                # id_papel = papel.upper()+ ':' + str(gramaje) + 'gr.-1.2 mano'
                id_papel = 'OFFSET BLANCO:80gr.-1.3 mano'
                paginas_bn = paginas - bitono - color
                print(estado, cod_interno, titulo, isbn, paginas, editorial, datovariable, id_papel,tamaño,solapas,bitono,color,paginas_bn)
                sql = "INSERT INTO bd_libro(estado, cod_interno, titulo, isbn, paginas, editorial, datovariable, id_papel,tamaño,solapas,bitono,color,paginas_bn) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                datos = (estado, cod_cliente, titulo, isbn, no_paginas, id_cliente, dato_variable,id_papel,id_formato,solapas,paginas_bitono,paginas_color, paginas_bn)
                cur.execute(sql, datos)
                conn.commit()
            except :
                
                print('error en el {}'.format(cod_interno))

        


        
        contador = contador + 1

    conn.close()



