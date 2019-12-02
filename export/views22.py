from django.shortcuts import render

from tablib import Dataset 

from .resources import Imp_LibroResource

from .models import Archivo
from .forms import FormEntrada
from bd.models import Libro
import openpyxl
import psycopg2
import datetime


# Create your views here.
def importar(request):
    if request.method == 'POST':
        form = FormEntrada(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['xlsfile']
            print(archivo)
            insert = Archivo(archivo = archivo)
            insert.save()
            traspaso(archivo)



    # template = loader.get_template('export/importar.html')  
    # if request.method == 'POST':
    #     libro_resource = Imp_LibroResource()  
    #     dataset = Dataset()
    #     print('dataset')  #
    #     print(dataset)  
    #     nuevos_libros = request.FILES['xlsfile']
    #     print('nuevos_libros')   
    #     print(nuevos_libros)  # NOMBRE DEL ARCHIVO
    #     imported_data = dataset.load(nuevos_libros.read())
    #     print('imported_data') 
    #     print(imported_data) # TODOS LOS DATOS
    #     result = libro_resource.import_data(dataset, dry_run=True) # Test the data import  
    #     print('result.has_errors')  
    #     print(result.has_errors())  
    #     if not result.has_errors():
    #         libro_resource.import_data(dataset, dry_run=False) # Actually import now  
    # return render(request, 'export/importar.html') 
   
            

        # conn = psycopg2.connect("host='localhost' port='5432' dbname='db_servinform' user='carlos' password='1234'")
        # cur = conn.cursor()
        # nuevos_libros = request.FILES['xlsfile']
        # print(nuevos_libros)
        # doc = openpyxl.load_workbook(nuevos_libros)
        # hojas = doc.sheetnames

        # hoja = doc['Sheet1']
        # contador = 1
        # contador2 = 1
        # for fila in hoja.rows:
        #     if contador <4:
        #         print(fila)
        #         contador = contador + 1
        #     else:
        #         break
    return render(request, 'export/importar.html')

def traspaso(archivo):
    print(archivo)
    ruta = 'media/' + str(archivo)
    doc = openpyxl.load_workbook(ruta)

    hojas = doc.sheetnames

    hoja = doc['Sheet1']
    contador = 1
    for fila in hoja.values:
        print(fila)
        
        if contador <4:
            for celda in fila:
                print(celda)
            print('')
            contador = contador + 1
        else:
            break
    
    

